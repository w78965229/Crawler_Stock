import os.path

import requests
from bs4 import BeautifulSoup
import datetime
import pandas as pd
from openpyxl import load_workbook
import  connect_bat as cb
import shutil #move file and rename file


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.support.ui import Select
#dealing with excel into db
import data_dealing as dd

# Stock_URL='https://www.twse.com.tw/zh/page/trading/exchange/STOCK_DAY.html'
Stock_List_URL='https://stock.wespai.com/lists'

Investment_Holding_Stock_URL='https://stock.wespai.com/p/63251'

OTC_URL='https://goodinfo.tw/tw/ShowK_Chart.asp?STOCK_ID=%E6%AB%83%E8%B2%B7%E6%8C%87%E6%95%B8&CHT_CAT2=DATE'
OSC_URL='https://goodinfo.tw/tw/ShowK_Chart.asp?STOCK_ID=%E5%8A%A0%E6%AC%8A%E6%8C%87%E6%95%B8&CHT_CAT2=DATE'
OSC_STOCK_List_URL='https://goodinfo.tw/tw2/StockList.asp?MARKET_CAT=%E4%B8%8A%E5%B8%82&INDUSTRY_CAT=%E4%B8%8A%E5%B8%82%E5%85%A8%E9%83%A8&SHEET=%E8%87%AA%E8%A8%82%E6%AC%84%E4%BD%8D_%E5%A4%96%E8%B3%87'
OTC_STOCK_List_URL='https://goodinfo.tw/tw2/StockList.asp?MARKET_CAT=%E4%B8%8A%E6%AB%83&INDUSTRY_CAT=%E4%B8%8A%E6%AB%83%E5%85%A8%E9%83%A8&SHEET=%E4%BA%A4%E6%98%93%E7%8B%80%E6%B3%81&SHEET2=%E6%97%A5&RPT_TIME=%E6%9C%80%E6%96%B0%E8%B3%87%E6%96%99'
# warnings.filterwarnings("ignore", category=UserWarning, module='bs4')
# -*- coding: utf-8 -*-
#coding=gbk
'''task1'''
def visit_Stock_URL(url):

    cookies = {
        "CLIENT%5FID": "20240508193926906%5F220%2E129%2E67%2E143",
        "IS_TOUCH_DEVICE": "F",
        "SCREEN_SIZE": "WIDTH=1920&HEIGHT=1080"
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
        'accept-encoding': 'gzip, deflate, br, zstd',
        'accept-language': 'zh-TW,zh;q=0.5',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'sec-ch-ua': '"Chromium";v="124", "Brave";v="124", "Not-A.Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'sec-gpc': '1',
        'upgrade-insecure-requests': '1',
        "Referer": "https://www.google.com/"
    }

    session = requests.Session()
    session.headers = headers
    response = session.get(url, cookies=cookies)
    content = response.content
    decoded_string = content.decode('utf-8')
    # print("decoded_string=",decoded_string)
    return decoded_string

def crawl_stock_info():
    week_day,exe_flg=cb.is_today_week_day_execute()
    print("week_day=",week_day," exe_flg=",exe_flg)

    '''test area-------------------'''
    # today = datetime.date.today().strftime('%Y/%m/%d')
    # OTC_html = visit_Stock_URL(OTC_URL)
    # OTC_today_info_list = get_OTC_today_info(OTC_html)
    # OSC_html = visit_Stock_URL(OSC_URL)
    # OSC_today_info_list = get_OSC_today_info(OSC_html)
    # return OTC_today_info_list,OSC_today_info_list
    '''----------------------------'''
    '''production'''
    if exe_flg==True:
        today = datetime.date.today().strftime('%Y/%m/%d')
        # print( week_day,exe_flg)

        '''上櫃指數OSC'''
        print("Get today OTC info")
        OTC_path = 'D:/PycharmProjects/milktea_project/OTC_OSC_info/K_Chart_OTC.xlsx'
        OTC_html = visit_Stock_URL(OTC_URL)
        # print("OTC_html=",OTC_html)

        OTC_today_info_list = get_OTC_today_info(OTC_html)
        if OTC_today_info_list !=[]:
            OTC_today_info_list.append("上櫃")
        # print("今日有上櫃指數有更新")
        print("今日上櫃指數LIST:", OTC_today_info_list)
        if cb.is_file_first_today(OTC_path, OTC_today_info_list[0],"OTC") == False:
            append_today_info_into_excel(OTC_today_info_list, OTC_path)


        '''上市指數OSC'''
        print("Get today OSC info")
        OSC_path = 'D:/PycharmProjects/milktea_project/OTC_OSC_info/K_Chart_OSC.xlsx'
        OSC_html = visit_Stock_URL(OSC_URL)
        OSC_today_info_list = get_OSC_today_info(OSC_html)

        if OSC_today_info_list !=[]:
            OSC_today_info_list.append("上市")
        # print("今日有上市指數有更新")
        print("今日上市指數LIST:", OSC_today_info_list)
        if cb.is_file_first_today(OSC_path, OSC_today_info_list[0],"OSC") == False:
            append_today_info_into_excel(OSC_today_info_list, OSC_path)



        if OTC_today_info_list and OSC_today_info_list:

            return OTC_today_info_list,OSC_today_info_list

    else:
        week_dict={0:'星期一',1:'星期二',2:'星期三',3:'星期四',4:'星期五',5:'星期六',6:'星期天'}

        print("today is "+str(week_dict[week_day]),'不需要執行程式')

        return None,None


def get_OTC_today_info(OTC_html):
    try:
        # print("get_OTC_today_info")
        soup = BeautifulSoup(OTC_html, 'html.parser')
        # print("get_OTC_today_info soup=",soup)

        center_route = soup.find('div', {"id": "divK_ChartDetail"}).find('div', {"id": "divPriceDetail"}).find('table',
                                                                                                               {
                                                                                                                   'id': 'tblPriceDetail'}).find_all(
            'tr')[2].find_all('td')
        # print(center_route)
        today = datetime.date.today().strftime('%Y/%m/%d')
        OTC_today_info_list = [i.text for i in center_route]
        OTC_today_info_list[0] = "20" + OTC_today_info_list[0][1:]
        # OTC_today_info_list[0]=datetime.datetime.strptime(today,'%Y-%m-%d')
        # '''test_area'''
        # return OTC_today_info_list
        '''production'''
        return OTC_today_info_list

        # if OTC_today_info_list[0] == today:
        #     return OTC_today_info_list
        # else:
        #     return []

        # path = 'D:\PycharmProjects\stock\stock_crawl_data\output.txt'
        # f = open(path, 'w',encoding='utf-8')
        # f.write(str(soup))
        # f.close()



    except Exception as e:
        print(e)
        return None


def get_OSC_today_info(OSC_html):
    try:
        # print("get_OTC_today_info")
        soup = BeautifulSoup(OSC_html, 'html.parser')
        # print("soup=")

        center_route = soup.find('div', {"id": "divK_ChartDetail"}).find('div', {"id": "divPriceDetail"}).find('table',
                                                                                                               {
                                                                                                                   'id': 'tblPriceDetail'}).find_all(
            'tr')[2].find_all('td')
        today = datetime.date.today().strftime('%Y-%m-%d')

        OSC_today_info_list = [i.text for i in center_route]
        OSC_today_info_list[0] = "20" + OSC_today_info_list[0][1:]
        # OTC_today_info_list[0]=datetime.datetime.strptime(today,'%Y-%m-%d')

        # '''test area'''
        '''production'''
        return OSC_today_info_list

        # if OSC_today_info_list[0] == today:
        #     return OSC_today_info_list
        # else:
        #     return []

        # print("OSC_today_info_list=",OSC_today_info_list)
        # path = 'D:\PycharmProjects\stock\stock_crawl_data\output.txt'
        # f = open(path, 'w',encoding='utf-8')
        # f.write(str(soup))
        # f.close()
    except Exception as e:
        print(e)
        return None

def append_today_info_into_excel(today_info,path):
    # today_info[0]= today_info[0].strftime("%Y/%m/%d") #資料轉換成YYYY/MM/DD
    wb = load_workbook(path)
    sheet = wb.worksheets[0]

    sheet.insert_rows(idx=2,amount=1)
    for i in range(len(today_info)) :
        sheet.cell(column=i+1, row=2).value = str(today_info[i])

    wb.save(path)

'''task2'''
'''檢查今天有多少excel是尚未下載的，每次執行只要下載沒下載過的就好'''
def check_download_excel(stock_type,category,today):
    print("check_download_excel start")
    new_file_root_path = 'D:\PycharmProjects\milktea_project\Excel_Data'
    foreign_investment_pipline_1 = ['法人買賣_外資']
    foreign_investment_pipline_2 = [
        ['法人買賣張數(日)', '法人買賣金額(百萬元)(日)', '法人買賣佔發行張數(日)', '法人買賣佔成交比重(日)','法人持股狀況(日)',
         '法人連買連賣統計(日)', '法人連買連賣轉折點(日)']
    ]

    investment_trust_pipline_1 = ['交易狀況','法人買賣_三大']
    investment_trust_pipline_2 = [
        ['日'],
        ['法人買賣張數(日)', '法人買賣金額(百萬元)(日)', '法人買賣佔發行張數(日)', '法人買賣佔成交比重(日)','法人持股狀況(日)',
         '法人連買連賣統計(日)', '法人連買連賣轉折點(日)']
    ]

    basic_info_pipline_1 = ['交易狀況', '融資融券', '季獲利能力', '年獲利能力', '近四季獲利能力', '營收狀況_近N個月一覽', '股利政策發放年度']
    basic_info_pipline_2 = [
        ['日'],
        ['資券增減統計(日)', '借券增減統計(日)'],
        ['獲利能力 (季增減統計)', '獲利能力 (年增減統計)'],
        ['獲利能力'],
        ['獲利能力 (季增減統計)', '獲利能力 (年增減統計)'],
        ['單月營收', '年增率'],
        ['股利分配資料 (以最後成交價統計)']
    ]

    technical_pipline_1 = ['交易狀況','移動均線', 'RSI', 'KD指標', 'MACD']
    technical_pipline_2 = [
        ['日'],
        ['目前位置1(元)', '乖離率1(%)'],
        ['未還原權值'],
        ['日', '日/週/月'],
        ['日/週/月', '季/年']
    ]
    pipline_1 = []
    pipline_2 = []

    not_ready_pipline_1 = []
    not_ready_pipline_2 = []
    # '季獲利能力', '年獲利能力', '近四季獲利能力'refresh後不用下載檔案 需先換第二個下拉式選單才可以下載檔案
    basic_list_without_download_1 = ['季獲利能力', '近四季獲利能力', '營收狀況_近N個月一覽']
    basic_list_without_download_2 = [['獲利能力 (季增減統計)'], ['獲利能力 (季增減統計)'], ['單月營收']]

    basic_list_without_download_dict ={
        '季獲利能力':['獲利能力 (季增減統計)'],
        '近四季獲利能力':['獲利能力 (季增減統計)'],
        '營收狀況_近N個月一覽':['單月營收']
    }

    if category == 'foreign_investment':
        pipline_1 = foreign_investment_pipline_1
        pipline_2 = foreign_investment_pipline_2

    elif category == 'investment_trust':
        pipline_1 = investment_trust_pipline_1
        pipline_2 = investment_trust_pipline_2

    elif category  == 'technical':
        pipline_1 = technical_pipline_1
        pipline_2 = technical_pipline_2
    elif category  == 'basic_info':
        pipline_1 = basic_info_pipline_1
        pipline_2 = basic_info_pipline_2
    check_exist_flag = True #True=要檢查，檢查若不存在要下載
    for i in range(len(pipline_1)):

        for j in range(len(pipline_2[i])):
            index=-1
            try:
                if basic_list_without_download_1.index(pipline_1[i]) >= 0:
                    index = basic_list_without_download_1.index(pipline_1[i])
                    # print(index)

                    if pipline_2[i][j] in basic_list_without_download_2[index]:
                        check_exist_flag=False
                    else:
                        check_exist_flag=True

            except ValueError:
                check_exist_flag=True
            # print("check_exist_flag=",check_exist_flag)
            if check_exist_flag:
                new_file_path = os.path.join(new_file_root_path, stock_type, category, today, transfer_filename_str(pipline_1[i]+'_'+pipline_2[i][j])+ '.xls')
                os.path.isfile(new_file_path)

                if os.path.isfile(new_file_path)==False :
                    if pipline_1[i] not in not_ready_pipline_1 : #20240731
                        not_ready_pipline_1.append(pipline_1[i])
                        not_ready_pipline_2.append(pipline_2[i])

                    print(f"The file '{new_file_path}' does not exist need to be download this time")
                # else:
                    # print(f"The file '{new_file_path}' exists")

    return not_ready_pipline_1,not_ready_pipline_2

def visit_by_google_account_Selenium(stock_type,input_url,today):

    try:

        # 配置Chrome选项
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--headless")

        # 初始化WebDriver

        driver = webdriver.Chrome()
        # 打开Google登录页面
        # OSC資料
        driver.get(input_url)
        # OTC_STOCK_List_URL
        time.sleep(20)

        # if download_list1 != [] :
        #     download_category_excel(driver,'foreign_investment')
        # driver.execute_script("arguments[0].click();", button)
        '''三大法人資料 (收集外資,投信.xls資源)'''
        download_list1,download_list2=check_download_excel(stock_type,'investment_trust',today)
        print("investment_trust download_list1=",download_list1)
        print("investment_trust download_list2=",download_list2)
        if download_list1!=[]:
            download_category_excel(driver,stock_type,'investment_trust',today,download_list1,download_list2)

        '''三大法人資料 (收集技術面excel資料)'''
        download_list1, download_list2 = check_download_excel(stock_type, 'technical',today)
        print("technical download_list1=", download_list1)
        print("technical download_list2=", download_list2)
        if download_list1 != []:
            download_category_excel(driver,stock_type,'technical',today,download_list1,download_list2)

        '''三大法人資料 (收集基本面excel資料)'''
        download_list1, download_list2 = check_download_excel(stock_type, 'basic_info',today)
        print("basic_info download_list1=", download_list1)
        print("basic_info download_list2=", download_list2)
        if download_list1 != []:
            download_category_excel(driver,stock_type,'basic_info',today,download_list1,download_list2)


        time.sleep(3)
        driver.quit()

        # 关闭浏览器 # 观察一下结果
        # time.sleep(10)

        # 创建requests会话并添加cookies
        # session = requests.Session()
        # for cookie in cookies:
        #     session.cookies.set(cookie['name'], cookie['value'])

        # 使用requests会话进行请求
        # response = session.get("https://www.example.com")
        # print(response.content)

    except Exception as e:
        print("An error occurred:", e)
        driver.quit()
def transfer_filename_str(new_file_name):

    if "/" in new_file_name:
        return new_file_name.replace("/", "_")
    return new_file_name
def download_category_excel(driver,stock_type,category,today,download_pipline_1,download_pipline_2):
    download_folder_path = "C:\\Users\\User\\Downloads"
    new_file_root_path = 'D:\PycharmProjects\milktea_project\Excel_Data'

    pipline=[]
    '''
    foreign_investment_pipline_1=['法人買賣_外資']
    foreign_investment_pipline_2=[
        ['法人買賣張數(日)','法人買賣金額(百萬元)(日)','法人買賣佔發行張數(日)','法人買賣佔成交比重(日)','法人持股狀況(日)',
         '法人連買連賣統計(日)','法人連買連賣轉折點(日)']
                                ]

    investment_trust_pipline_1 = ['法人買賣_三大']

    investment_trust_pipline_2=[
        ['法人買賣張數(日)','法人買賣金額(百萬元)(日)','法人買賣佔發行張數(日)','法人買賣佔成交比重(日)','法人持股狀況(日)',
         '法人連買連賣統計(日)','法人連買連賣轉折點(日)']
                                ]

    basic_info_pipline_1   = ['交易狀況','融資融券','季獲利能力','年獲利能力','近四季獲利能力','營收狀況_近N個月一覽','股利政策發放年度']
    basic_info_pipline_2   = [
                              ['日'],
                              ['資券增減統計(日)','借券增減統計(日)'],
                              ['獲利能力 (季增減統計)','獲利能力 (年增減統計)'],
                              ['獲利能力'],
                              ['獲利能力 (季增減統計)', '獲利能力 (年增減統計)'],
                              ['單月營收','年增率'],
                              ['股利分配資料 (以最後成交價統計)']
                              ]


    technical_pipline_1   = ['移動均線','RSI','KD指標','MACD']
    technical_pipline_2   = [
                             ['目前位置1(元)','乖離率1(%)'],
                             ['未還原權值'],
                             ['日','日/週/月'],
                             ['日/週/月','季/年']
                                                            ]
    pipline_selsheet1 = []
    pipline_selsheet2 = []




    if category == 'foreign_investment':
        pipline_selsheet1 = foreign_investment_pipline_1
    elif category == 'investment_trust':
        pipline_selsheet1 = investment_trust_pipline_1
    elif category  == 'technical':
        pipline_selsheet1 = technical_pipline_1
    elif category  == 'basic_info':
        pipline_selsheet1 = basic_info_pipline_1
'''
    '''setting file'''
    # '季獲利能力', '年獲利能力', '近四季獲利能力'refresh後不用下載檔案 需先換第二個下拉式選單才可以下載檔案
    basic_list_without_download = ['季獲利能力','近四季獲利能力','營收狀況_近N個月一覽']


    long_time_list_sel1sheet = ['營收狀況_近N個月一覽','股利政策發放年度']
    long_time_list_sel2sheet = ['年增率']
    pipline_selsheet1=download_pipline_1

    try:
        for k in range(0,len(pipline_selsheet1)):
            pipline_selsheet2 = download_pipline_2[k]

            # if category == 'foreign_investment':
                # pipline_selsheet2 = foreign_investment_pipline_2[k]
            # elif category == 'investment_trust':
            #     pipline_selsheet2 = download_pipline_2[k]
            # elif category == 'technical':
            #     pipline_selsheet2 = download_pipline_2[k]
            # elif category == 'basic_info':
            #     pipline_selsheet2 = download_pipline_2[k]

            '''下載第一次:法人買賣張數(日)的EXCEL'''
            selSHEET = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "selSHEET"))
            )
            select = Select(selSHEET)
            # 通过可见文本选择选项
            select.select_by_value(pipline_selsheet1[k])

            # 验证选择是否成功
            selected_option = select.first_selected_option
            print(selected_option)
            print("Selected option:", selected_option.text)

            if pipline_selsheet1[k] in long_time_list_sel1sheet:
                time.sleep(20)  # 等待頁面refresh
            else:
                time.sleep(10)  # 等待頁面refresh
            # 抓取export按鈕
            button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[value='匯出XLS']"))  # 使用ID定位按钮
            )
            print(button.get_attribute('outerHTML'))

            #有些資料不是selsheet1 refresh後就需要下載
            if pipline_selsheet1[k] not in basic_list_without_download:
                # 点选按钮
                button.click()
                '''處理更改下載後的檔名並將其移走'''
                time.sleep(2)  # 下載後等待兩秒
                rename_and_move_file(new_file_root_path,download_folder_path,stock_type,category,today,transfer_filename_str(pipline_selsheet1[k]+"_"+pipline_selsheet2[0]))

            if len(pipline_selsheet2)>1:
                for i in range(1,len(pipline_selsheet2)):

                    # 通过可见文本选择选项
                    time.sleep(2)  # 下載後等待兩秒
                    selSHEET2 = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "selSHEET2"))
                    )
                    select = Select(selSHEET2)
                    select.select_by_value(pipline_selsheet2[i])
                    selected_option = select.first_selected_option
                    print('selSHEET2.option=',selected_option.text)
                    if pipline_selsheet2[i] in long_time_list_sel2sheet:

                        time.sleep(30)  # 等待頁面refresh
                    else:
                        time.sleep(10)  # 等待頁面refresh

                    for j in range(1,10):
                        try:
                            print('this is j times=',j)
                            button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[value='匯出XLS']"))  # 使用ID定位按钮
                            )

                            print(button.get_attribute('outerHTML'))
                            # 点选按钮
                            button.click()

                            time.sleep(2)  # 下載後等待兩秒
                            rename_and_move_file(new_file_root_path, download_folder_path, stock_type,category, today, transfer_filename_str(pipline_selsheet1[k]+"_"+pipline_selsheet2[i]))

                            break
                        except Exception as e:
                            print(f"An error occurred during click: {e}")
                            # 重新查找按钮元素
                            ad_close_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@class='ad-close-button']"))
                            )
                            ad_close_button.click()

                            WebDriverWait(driver, 10).until_not(
                                EC.presence_of_element_located((By.XPATH, "//div[@class='ad-container']"))
                            )

                            print("exception try button :", button.get_attribute('outerHTML'))
                            time.sleep(1)  # 等待1秒后重试






    except Exception as e:
        print("An error occurred:", e)

        ad_close_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='ad-close-button']"))
        )
        ad_close_button.click()

        WebDriverWait(driver, 10).until_not(
            EC.presence_of_element_located((By.XPATH, "//div[@class='ad-container']"))
        )

        driver.quit()

def rename_and_move_file( root_path, destination_dir,stock_type,category,today,new_file_name):
    try:
        #更名後的excel檔案會放在 target_path:root_path+"\"+category+"\"+today+"\"+new_file_name+".xlsx"
        # 确保目标目录存在，不存在则创建

        if not os.path.exists(root_path):
            os.makedirs(root_path)

        if not os.path.exists(os.path.join(root_path,stock_type)):
            os.makedirs(os.path.join(root_path,stock_type))

        if not os.path.exists(os.path.join(root_path,stock_type,category)):
            os.makedirs(os.path.join(root_path,stock_type,category))

        if not os.path.exists(os.path.join(root_path, stock_type,category,today)):
            os.makedirs(os.path.join(root_path, stock_type,category,today))


        # if not os.path.exists(destination_dir):
        #     os.makedirs(destination_dir)

        # 获取源文件的目录
        source_file = os.path.join(destination_dir,'StockList.xls')

        # print("source_file=",source_file)
        # 新的文件路径
        new_file_path = os.path.join(root_path, stock_type,category,today,new_file_name+'.xls')
        # print("new_file_path=",new_file_path)

        # 重命名并移动文件
        shutil.move(source_file, new_file_path)
        print(f" File From {source_file} renamed and moved to: {new_file_path}")


    except Exception as e:
        print(f"An error occurred: {e}")


def get_stock_info(stock_type_list,dealing_list): #抓取個股資訊
    week_day,exe_flg=cb.is_today_week_day_execute()
    print("week_day=",week_day," exe_flg=",exe_flg)
    today = datetime.date.today().strftime('%Y%m%d')
    # today ='20240731'
    '''production'''
    exe_flg=True
    if exe_flg==True:

        # '''上市相關excel資料'''
        visit_by_google_account_Selenium('OSC_上市',OSC_STOCK_List_URL,today)
        # '''上櫃相關excel資料'''
        visit_by_google_account_Selenium('OTC_上櫃',OTC_STOCK_List_URL,today)

        '''dealing with excel to database'''


        root_path='D:\PycharmProjects\milktea_project\Excel_Data'
        stock_dataframe_dict = {}
        for stock_type in stock_type_list:
            stock_dataframe_dict[stock_type] = {}

            # stock_dataframe_dict[stock_type] = pd.DataFrame()

            for deal in dealing_list:
                print(f"正在整理 '{stock_type}'的 {deal} 資料 ")
                stock_dataframe_dict[stock_type][deal]=pd.DataFrame()
                root_file_path = os.path.join(root_path,stock_type,deal,today)
                stock_dataframe_dict=dd.dealing_excel_into_dataframe(root_file_path,stock_type,deal,today,stock_dataframe_dict)

        print("stock_dataframe_dict.keys()=",list(stock_dataframe_dict))
        print("stock_dataframe_dict=",stock_dataframe_dict)


        return stock_dataframe_dict
        # '''各股資訊'''
        # print("Get today stocks info 抓取個股資訊")
        # stock_list_html_t = visit_Stock_URL(STOCK_List_URL_t)
        # path = 'D:/PycharmProjects/milktea_project/test_area/task2_parse.txt'
        # f = open(path, 'w',encoding='utf-8')
        # f.write(stock_list_html_t)
        # f.close()

        # print("stock_list_html_t=",stock_list_html_t)
        # get_stock_list(stock_list_html_t)
    else:
        week_dict = {0: '星期一', 1: '星期二', 2: '星期三', 3: '星期四', 4: '星期五', 5: '星期六', 6: '星期天'}

        print("today is " + str(week_dict[week_day]), '不需要執行程式')
        return None
def get_inverst_holding_rate(URL_content,today):
    try:
        soup = BeautifulSoup(URL_content, 'html.parser')
        center_route = soup.find('body').find('div', {"class": "ad"}).find("table",   {"class": "display", "id": "example"})
        header_list_soup = center_route.find("thead").find("tr").find_all('th')

        header_list = ['日期']
        for i in range(len(header_list_soup)):
            header_list.append(header_list_soup[i].text)
        # print("length of header_list=",len(header_list))

        body_list_soup = center_route.find("tbody").find_all("tr")
        body_list = []
        for i in range(len(body_list_soup)):
            temp_body_list = [today]
            # temp_body_list = ['2022-07-22']
            body_stock_info_list_soup = body_list_soup[i].find_all('td')
            for j in range(len(body_stock_info_list_soup)):
                temp_body_list.append(body_stock_info_list_soup[j].text)
            body_list.append(temp_body_list)
        # print("header_list=",header_list)
        # print("body_list=",body_list)

        stock_df = pd.DataFrame(columns=header_list)
        for i in range(len(body_list)):
            stock_df.loc[i] = body_list[i]


        file_name = "Holding_Rate_" + today + ".xlsx"
        file_folder_path = "D:/PycharmProjects/milktea_project/Excel_Data/Holding_Rate"

        if not os.path.exists(file_folder_path):
            os.makedirs(file_folder_path)

        if os.path.exists(os.path.join(file_folder_path, file_name)):
            print(file_name + " is exists in " + file_folder_path)
        else:
            print(file_name + " is not exists in " + file_folder_path + " Creating the Excel..in it")

            stock_df.to_excel(os.path.join(file_folder_path, file_name), index=False)
        # .find("div",{"class":"dataTables_scroll"}).find("div", {"class": "dataTables_scrollBody"}).find("table", {"id": "example"}).find("tbody").find_all('tr')
        # print(div_frame)
        return stock_df
    except Exception as e:
            print(e)
            return None


def visit_Holding_rate_URL():
    data = {
        'start': 20,
        'limit': 20,
        'sort': 'time',
        'status': 'P'
    }
    user_agent = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.109 Safari/537.36"
    response = requests.get(Investment_Holding_Stock_URL, params=data, headers={'User-Agent': user_agent})
    response.raise_for_status()  # 如果返回的状态码不是200， 则抛出异常;
    # response.encoding = response.apparent_encoding  # 判断网页的编码格式，便于respons.text知道如何解码;
    response.encoding = 'utf-8'
    print("response=",response.status_code)
    if response.status_code!=200:

        print('Invalid url:',response.url)
        return None
    else:


        return response.text